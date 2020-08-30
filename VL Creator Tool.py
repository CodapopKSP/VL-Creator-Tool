
# === Setup =======================================================================
import tkinter as tk
from tkinter import StringVar, IntVar
import csv
from datetime import timedelta, date
from openpyxl import Workbook
import ast
from openpyxl.styles import Font, Alignment, Border, Side

root = tk.Tk()
root.title("VL Creator Tool")
root.resizable(False, False)
root.geometry("500x400+2500+300")

headerFrame = tk.Frame(root)
bookSelect = tk.Frame(root, width=500, highlightbackground="purple", highlightthickness=1)
book1 = tk.Frame(bookSelect, bg="#E5E7E9", width=100, highlightbackground="purple", highlightthickness=1)
book2 = tk.Frame(bookSelect, bg="#ECF0F1", width=100, highlightbackground="purple", highlightthickness=1)
book3 = tk.Frame(bookSelect, bg="#E5E7E9", width=100, highlightbackground="purple", highlightthickness=1)
book4 = tk.Frame(bookSelect, bg="#ECF0F1", width=100, highlightbackground="purple", highlightthickness=1)
book5 = tk.Frame(bookSelect, bg="#E5E7E9", width=100, highlightbackground="purple", highlightthickness=1)

wordSelectCanvas = tk.Canvas(root, bg="white", width=473)
wordSelectFrame = tk.Frame(wordSelectCanvas, bg="white", width=473)
vscrollbar = tk.Scrollbar(root, orient="vertical", command=wordSelectCanvas.yview)
hscrollbar = tk.Scrollbar(root, orient="horizontal", command=wordSelectCanvas.xview)
wordSelectCanvas.configure(yscrollcommand=vscrollbar.set)
wordSelectCanvas.configure(xscrollcommand=hscrollbar.set)
#-----------------------------------------------------------------------------------

books = ["SBS 1", "Phonics A-2", "Reading A", "Write Now 1"]
files = ['Books/Side By Side/SBS 1.csv', 'Books/Phonics/Phonics A-2.csv', 'Books/Reading/Reading A.csv', 'Books/Writing/Write Now 1.csv']

# === Retrieve Entries from Files ==================================================
def submitBook(bookName, textA, textB):
	for i in range(len(books)):
		if (bookName == books[i]):
			with open(files[i]) as csv_file:
				csv_reader = csv.reader(csv_file, delimiter=';')
				for row in csv_reader:
					if ((int(row[0]) >= int(textA)-1) and (int(row[0]) <= int(textB))):
						print(row)
						with open('dataHandler.csv', mode='a', newline="") as dataHandler:
							dataHandler_write = csv.writer(dataHandler, delimiter=';', quotechar='"', quoting=csv.QUOTE_MINIMAL)
							dataHandler_write.writerow([row[1], row[2], row[3], row[4]])
#-----------------------------------------------------------------------------------


# === Book Select Functions ================================================================
def selectBooks():
	clearDisplay()
	with open('dataHandler.csv', mode='w', newline="") as dataHandler:
		dataHandler_write = csv.writer(dataHandler, delimiter=';', quotechar='"', quoting=csv.QUOTE_MINIMAL)
		dataHandler_write.writerow([])
	submitBook(readDrop1(), readText1a(), readText1b())
	submitBook(readDrop2(), readText2a(), readText2b())
	submitBook(readDrop3(), readText3a(), readText3b())
	submitBook(readDrop4(), readText4a(), readText4b())
	submitBook(readDrop5(), readText5a(), readText5b())
	displayData()
	# --- Update Status Bar ---
	bookStats1 = readDrop1() + ": p" + readText1a() + "-" + readText1b()
	bookStats2 = readDrop2() + ": p" + readText2a() + "-" + readText2b()
	bookStats3 = readDrop3() + ": p" + readText3a() + "-" + readText3b()
	bookStats4 = readDrop4() + ": p" + readText4a() + "-" + readText4b()
	bookStats5 = readDrop5() + ": p" + readText5a() + "-" + readText5b()
	statusText.set(bookStats1 + "  |  " + bookStats2 + "  |  " + bookStats3 + "  |  " + bookStats4 + "  |  " + bookStats5)
	
# ---Book 1---
def readDrop1():
	drop1read = dropVal1.get()
	return drop1read
def readText1a():
	text1a = entry1a.get("1.0", "end-1c")
	return text1a
def readText1b():
	text1b = entry1b.get("1.0", "end-1c")
	return text1b
# --- Book 2---
def readDrop2():
	drop2read = dropVal2.get()
	return drop2read
def readText2a():
	print(entry2a.get("1.0", "end-1c"))
	text2a = entry2a.get("1.0", "end-1c")
	return text2a
def readText2b():
	text2b = entry2b.get("1.0", "end-1c")
	return text2b
# ---Book 3---
def readDrop3():
	drop3read = dropVal3.get()
	return drop3read
def readText3a():
	text3a = entry3a.get("1.0", "end-1c")
	return text3a
def readText3b():
	text3b = entry3b.get("1.0", "end-1c")
	return text3b
# ---Book 4---
def readDrop4():
	drop4read = dropVal4.get()
	return drop4read
def readText4a():
	text4a = entry4a.get("1.0", "end-1c")
	return text4a
def readText4b():
	text4b = entry4b.get("1.0", "end-1c")
	return text4b
# ---Book 5---
def readDrop5():
	drop5read = dropVal5.get()
	return drop5read
def readText5a():
	text5a = entry5a.get("1.0", "end-1c")
	return text5a
def readText5b():
	text5b = entry5b.get("1.0", "end-1c")
	return text5b
#-----------------------------------------------------------------------------------
words = []
var = []

def displayData():
	buttonBuildVL.pack(side="bottom", pady=10)
	vscrollbar.pack(side="right", fill="y")
	hscrollbar.pack(side="bottom", anchor="w", fill="x")
	wordSelectCanvas.pack(side="right", anchor="w", pady=10, fill="x")
	wordSelectCanvas.create_window((0, 0), window=wordSelectFrame, anchor='n')
	wordSelectFrame.bind("<Configure>", scroll_CB)
	with open('dataHandler.csv', mode='r') as dataHandler:
		data=list(csv.reader(dataHandler, delimiter=';', quotechar='|'))
		for row in data:
			if (len(row)>0):
				word = StringVar()
				checkVar = tk.BooleanVar()
				word.set(row)
				words.append(row)
				var.append(checkVar)
				display = tk.Checkbutton(wordSelectFrame, variable=checkVar, textvariable=word, fg="blue", bg="white", bd=1, relief="sunken")
				display.pack(side="top", anchor="w")	

def clearDisplay():
	for display in wordSelectFrame.winfo_children():
		display.destroy()
		words.clear()
		var.clear()

def scroll_CB(event):
	wordSelectCanvas.configure(scrollregion=wordSelectCanvas.bbox("all"))
	wordSelectCanvas.xview_moveto(0)

def set_border(ws, cell_range):
    border = Border(left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000'))
    rows = ws[cell_range]
    for row in rows:
        for cell in row:
            cell.border = border

def buildVL():
	workbook = Workbook()
	vocabList = workbook.active

	# --- Build the Template ---
	vocabList.merge_cells('C1:D1')
	vocabList.column_dimensions['A'].width = 2
	vocabList.column_dimensions['B'].width = 12
	vocabList.column_dimensions['C'].width = 12
	vocabList.column_dimensions['D'].width = 24
	vocabList.column_dimensions['E'].width = 35

	levelNumber = levelEntry.get("1.0", "end-1c")
	weekNumber = weekEntry.get("1.0", "end-1c")
	vocabList["B1"] = ("Level " + levelNumber)
	vocabList["B1"].font = Font(name="Arial", size=12, bold=True)
	vocabList["C1"] = ("Vocabulary Week " + weekNumber)
	vocabList["C1"].font = Font(name="Arial", size=12, bold=True)
	vocabList["C1"].alignment = Alignment(horizontal="center")
	vocabList["E1"] = "Name: ______________"
	vocabList["E1"].font = Font(name="Arial", size=12, bold=True)
	vocabList["E1"].alignment = Alignment(horizontal="center")
	dayCheck = 0
	dayCell = 2
	if monVar.get() == 1:
		dayCheck = dayCheck + 1
		vocabList["B2"] = "Monday"
		vocabList["B" + str(dayCell)].font = Font(name="Arial", size=12, bold=True)
		set_border(vocabList, "A" + str(dayCell + 1) + ":" + "E" + str(dayCell + 6))
		if autoDate.get() == 1:
			vocabList["E2"] = str(nextMonday)
			vocabList["E" + str(dayCell)].font = Font(name="Arial", size=12, bold=True)
			vocabList["E" + str(dayCell)].alignment = Alignment(horizontal="right")
		dayCell = dayCell + 7

	if tueVar.get() == 1:
		dayCheck = dayCheck + 1
		vocabList["B" + str(dayCell)] = "Tuesday"
		vocabList["B" + str(dayCell)].font = Font(name="Arial", size=12, bold=True)
		set_border(vocabList, "A" + str(dayCell + 1) + ":" + "E" + str(dayCell + 6))
		if autoDate.get() == 1:
			vocabList["E" + str(dayCell)] = str(nextTuesday)
			vocabList["E" + str(dayCell)].font = Font(name="Arial", size=12, bold=True)
			vocabList["E" + str(dayCell)].alignment = Alignment(horizontal="right")
		dayCell = dayCell + 7

	if wedVar.get() == 1:
		dayCheck = dayCheck + 1
		vocabList["B" + str(dayCell)] = "Wednesday"
		vocabList["B" + str(dayCell)].font = Font(name="Arial", size=12, bold=True)
		set_border(vocabList, "A" + str(dayCell + 1) + ":" + "E" + str(dayCell + 6))
		if autoDate.get() == 1:
			vocabList["E" + str(dayCell)] = str(nextWednesday)
			vocabList["E" + str(dayCell)].font = Font(name="Arial", size=12, bold=True)
			vocabList["E" + str(dayCell)].alignment = Alignment(horizontal="right")
		dayCell = dayCell + 7

	if thuVar.get() == 1:
		dayCheck = dayCheck + 1
		vocabList["B" + str(dayCell)] = "Thursday"
		vocabList["B" + str(dayCell)].font = Font(name="Arial", size=12, bold=True)
		set_border(vocabList, "A" + str(dayCell + 1) + ":" + "E" + str(dayCell + 6))
		if autoDate.get() == 1:
			vocabList["E" + str(dayCell)] = str(nextThursday)
			vocabList["E" + str(dayCell)].font = Font(name="Arial", size=12, bold=True)
			vocabList["E" + str(dayCell)].alignment = Alignment(horizontal="right")
		dayCell = dayCell + 7

	rowCount = 3
	printCount = 0
	while printCount < dayCheck:
		vocabList[("B" + str(rowCount))] = "Word"
		vocabList[("B" + str(rowCount))].font = Font(name="Arial", size=12, bold=True)
		vocabList[("C" + str(rowCount))] = "Syllables"
		vocabList[("C" + str(rowCount))].font = Font(name="Arial", size=12, bold=True)
		vocabList[("D" + str(rowCount))] = "Definition"
		vocabList[("D" + str(rowCount))].font = Font(name="Arial", size=12, bold=True)
		vocabList[("E" + str(rowCount))] = "Sentence"
		vocabList[("E" + str(rowCount))].font = Font(name="Arial", size=12, bold=True)
		vocabList[("A" + str(rowCount + 1))] = "1"
		vocabList[("A" + str(rowCount + 2))] = "2"
		vocabList[("A" + str(rowCount + 3))] = "3"
		vocabList[("A" + str(rowCount + 4))] = "4"
		vocabList[("A" + str(rowCount + 5))] = "5"
		vocabList.row_dimensions[rowCount + 1].height = 25
		vocabList.row_dimensions[rowCount + 2].height = 25
		vocabList.row_dimensions[rowCount + 3].height = 25
		vocabList.row_dimensions[rowCount + 4].height = 25
		vocabList.row_dimensions[rowCount + 5].height = 25
		vocabList[("A" + str(rowCount + 1))].font = Font(name="Arial", size=12, bold=True)
		vocabList[("A" + str(rowCount + 2))].font = Font(name="Arial", size=12, bold=True)
		vocabList[("A" + str(rowCount + 3))].font = Font(name="Arial", size=12, bold=True)
		vocabList[("A" + str(rowCount + 4))].font = Font(name="Arial", size=12, bold=True)
		vocabList[("A" + str(rowCount + 5))].font = Font(name="Arial", size=12, bold=True)
		rowCount = rowCount + 7
		printCount = printCount + 1

		# --- Get Words/Sentences ---
	displayCount = -1
	columnCount = 0
	wordRowCount = 4
	dayCheckMulti = (dayCheck * 7) + 4
	for display in wordSelectFrame.winfo_children():
		displayCount=displayCount+1
		if var[displayCount].get() == 1:
			list1 = str(words[displayCount])
			list1 = ast.literal_eval(list1)
			vocabList["B" + str(wordRowCount)] = list1[columnCount]
			vocabList["C" + str(wordRowCount)] = list1[columnCount+1]
			vocabList["D" + str(wordRowCount)] = list1[columnCount+2]
			vocabList["E" + str(wordRowCount)] = list1[columnCount+3]
			vocabList["B" + str(wordRowCount)].font = Alignment(wrapText=True, vertical="center")
			vocabList["C" + str(wordRowCount)].font = Alignment(wrapText=True, vertical="center")
			vocabList["D" + str(wordRowCount)].font = Alignment(wrapText=True, vertical="center")
			vocabList["E" + str(wordRowCount)].font = Alignment(wrapText=True, vertical="center")
			vocabList["B" + str(wordRowCount)].font = Font(name="Arial", size=12)
			vocabList["C" + str(wordRowCount)].font = Font(name="Arial", size=12)
			vocabList["D" + str(wordRowCount)].font = Font(name="Arial", size=12)
			vocabList["E" + str(wordRowCount)].font = Font(name="Arial", size=12)
			wordRowCount = wordRowCount + 1
			if (wordRowCount==9):
				wordRowCount = wordRowCount + 2
			if (wordRowCount==16):
				wordRowCount = wordRowCount + 2
			if (wordRowCount==23):
				wordRowCount = wordRowCount + 2
		workbook.save(filename="Output/Vocab List Week " + str(weekNumber) + ".xlsx")
	dayCheck = 0




# === Header Parameters =================================================================
headerFrame.pack(side="top", pady=10)
weekEntryLabel = tk.Label(headerFrame, text="Week:")
weekEntry = tk.Text(headerFrame, height=1, width=3)
levelEntryLabel = tk.Label(headerFrame, text="Level:")
levelEntry = tk.Text(headerFrame, height=1, width=3)
monVar = tk.BooleanVar()
tueVar = tk.BooleanVar()
wedVar = tk.BooleanVar()
thuVar = tk.BooleanVar()
autoDate = tk.BooleanVar()
mondayCheckBox = tk.Checkbutton(headerFrame, text="Mon", variable=monVar, bg="#E5E7E9")
tuesdayCheckBox = tk.Checkbutton(headerFrame, text="Tue", variable=tueVar, bg="#E5E7E9")
wednesdayCheckBox = tk.Checkbutton(headerFrame, text="Wed", variable=wedVar, bg="#E5E7E9")
thursdayCheckBox = tk.Checkbutton(headerFrame, text="Thurs", variable=thuVar, bg="#E5E7E9")
dateCheckBox = tk.Checkbutton(headerFrame, text="Use Auto Dates", variable=autoDate)

weekEntryLabel.grid(row=0, column=0, sticky="e")
weekEntry.grid(row=0, column=1, sticky="w", padx=5)
levelEntryLabel.grid(row=0, column=2, sticky="e")
levelEntry.grid(row=0, column=3, sticky="w", padx=5)

mondayCheckBox.grid(row=0, column=4)
tuesdayCheckBox.grid(row=0, column=5)
wednesdayCheckBox.grid(row=0, column=6)
thursdayCheckBox.grid(row=0, column=7)
mondayCheckBox.select()
tuesdayCheckBox.select()
wednesdayCheckBox.select()
thursdayCheckBox.select()
dateCheckBox.grid(row=0, column=8)
#-----------------------------------------------------------------------------------


# === Date Management ==============================================================
today = date.today()
dayOfWeek = today.weekday()
daysUntilSunday = dayOfWeek - 7
dayCount = 0
if dayOfWeek > 0:
	while daysUntilSunday < 0:
		dayCount = dayCount + 1
		daysUntilSunday = daysUntilSunday + 1
		if daysUntilSunday == 0:
			break
nextMonday = (today + timedelta(days=dayCount))
nextTuesday = (today + timedelta(days=dayCount + 1))
nextWednesday = (today + timedelta(days=dayCount + 2))
nextThursday = (today + timedelta(days=dayCount + 3))
#-----------------------------------------------------------------------------------


# === Book Select Parameters =======================================================
bookTitle1 = tk.Label(book1, text="Book 1", bg="#E5E7E9", fg="purple")
bookTitle2 = tk.Label(book2, text="Book 2", bg="#ECF0F1", fg="purple")
bookTitle3 = tk.Label(book3, text="Book 3", bg="#E5E7E9", fg="purple")
bookTitle4 = tk.Label(book4, text="Book 4", bg="#ECF0F1", fg="purple")
bookTitle5 = tk.Label(book5, text="Book 5", bg="#E5E7E9", fg="purple")

# ---Drop Menus---
dropVal1 = StringVar()
dropVal2 = StringVar()
dropVal3 = StringVar()
dropVal4 = StringVar()
dropVal5 = StringVar()
drop1 = tk.OptionMenu(book1, dropVal1, *books)
drop2 = tk.OptionMenu(book2, dropVal2, *books)
drop3 = tk.OptionMenu(book3, dropVal3, *books)
drop4 = tk.OptionMenu(book4, dropVal4, *books)
drop5 = tk.OptionMenu(book5, dropVal5, *books)
# ---Page Number Entries---
entry1a = tk.Text(book1, height=1, width=4)
entry1b = tk.Text(book1, height=1, width=4)
entry2a = tk.Text(book2, height=1, width=4)
entry2b = tk.Text(book2, height=1, width=4)
entry3a = tk.Text(book3, height=1, width=4)
entry3b = tk.Text(book3, height=1, width=4)
entry4a = tk.Text(book4, height=1, width=4)
entry4b = tk.Text(book4, height=1, width=4)
entry5a = tk.Text(book5, height=1, width=4)
entry5b = tk.Text(book5, height=1, width=4)
#-----------------------------------------------------------------------------------

# === Book Select Bar ==============================================================
bookSelect.pack(side="top", pady=10)
# ---Book 1---
book1.grid(row=0, column=0)
bookTitle1.grid(row=0, column=0, columnspan=2)
drop1.grid(row=1, column=0, columnspan=2)
entry1a.grid(row=2, column=0, padx=5)
entry1b.grid(row=2, column=1, padx=5)
# ---Book 2---
book2.grid(row=0, column=1)
bookTitle2.grid(row=0, column=0, columnspan=2)
drop2.grid(row=1, column=0, columnspan=2)
entry2a.grid(row=2, column=0, padx=5)
entry2b.grid(row=2, column=1, padx=5)
# ---Book 3---
book3.grid(row=0, column=2)
bookTitle3.grid(row=0, column=0, columnspan=2)
drop3.grid(row=1, column=0, columnspan=2)
entry3a.grid(row=2, column=0, padx=5)
entry3b.grid(row=2, column=1, padx=5)
# ---Book 4---
book4.grid(row=0, column=3)
bookTitle4.grid(row=0, column=0, columnspan=2)
drop4.grid(row=1, column=0, columnspan=2)
entry4a.grid(row=2, column=0, padx=5)
entry4b.grid(row=2, column=1, padx=5)
# ---Book 5---
book5.grid(row=0, column=4)
bookTitle5.grid(row=0, column=0, columnspan=2)
drop5.grid(row=1, column=0, columnspan=2)
entry5a.grid(row=2, column=0, padx=5)
entry5b.grid(row=2, column=1, padx=5)
#-----------------------------------------------------------------------------------


# === Submit ======================================================================
buttonSubmit = tk.Button(root, width=30, fg="purple", text="Submit", command=selectBooks)
buttonSubmit.pack(side="top", pady=10)

# === Status Bar ===
statusText = StringVar(root)
statusText.set('Select books and pages...')
statusBar = tk.Label(root, textvariable=statusText, fg="blue", bd=1, relief="sunken", anchor="w")
statusBar.pack(side="bottom", fill="x")

# === Build Bar ===
buttonBuildVL = tk.Button(root, width=30, fg="blue", text="Build!", command=buildVL)
#-----------------------------------------------------------------------------------



# === Run It! ======================================================================
root.mainloop()
#-----------------------------------------------------------------------------------